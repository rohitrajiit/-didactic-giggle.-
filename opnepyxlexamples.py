import openpyxl
wb = openpyxl.load_workbook('example.xlsx')
print(wb.get_sheet_names())
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet.title)
anotherSheet = wb.get_active_sheet()
print(anotherSheet)
sheet = wb.get_sheet_by_name('Sheet3')
print(sheet.title)
sheet = wb.active
print(sheet)
print(sheet['A1'])
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
d='Row ' + str(c.row) + ', Column ' + c.column + ' is ' + c.value
print(d)
print(sheet.cell(row=1, column=2).value)
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
print(sheet.get_highest_row())
print(sheet.get_highest_column())
print(sheet.max_row)
print(sheet.max_column)
from openpyxl.cell import get_column_letter, column_index_from_string
print(get_column_letter(1))
print(get_column_letter(sheet.max_row))
print(tuple(sheet['A1':'C3']))
for rowOfCellObjects in sheet['A1':'C3']:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print('--- END OF ROW ---')
for cellObj in sheet.columns[1]:
    print(cellObj.value)
for cellObj in sheet.rows[1]:
    print(cellObj.value)
sheet = wb.active
sheet.title = 'Spam Bacon Eggs Sheet'
print( sheet.title)
wb.create_sheet(index=2, title='Middle Sheet')
print(wb.get_sheet_names())
wb.remove_sheet(wb.get_sheet_by_name('Middle Sheet'))
print(wb.get_sheet_names())
a= sheet['A1'].value
sheet['A1'] = 'Hello world!'
print('sheet a1 value',sheet['A1'].value)
sheet['A1'] = a
wb.save('example_copy.xlsx')
