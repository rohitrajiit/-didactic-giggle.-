import openpyxl
wb = openpyxl.Workbook()
sheet = wb.active
sheet['A1'] = 200
sheet['A2'] = 300
sheet['A3'] = '=SUM(A1:A2)'
wb.save('writeFormula.xlsx')
wbFormulas = openpyxl.load_workbook('writeFormula.xlsx')
sheet = wbFormulas.active
print(sheet['A3'].value)

wbDataOnly = openpyxl.load_workbook('writeFormula.xlsx', data_only=True)
sheet = wbDataOnly.active
print(sheet['A3'].value)
